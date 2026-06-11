"""app/services/jamil/extractor.py — Extract text content from any uploaded file.

Supports: PDF, DOCX, XLSX/CSV, images (via vision LLM), audio/video (via Groq Whisper),
plain text, code files, JSON, XML, and generic binary fallback.
"""
import io
import logging
import os

logger = logging.getLogger("lcewai")

# Max characters to include from a single file in context
_MAX_CHARS = 12000


def _ext(filename: str) -> str:
    return (filename or "").rsplit(".", 1)[-1].lower()


async def extract(filename: str, content: bytes, content_type: str) -> str:
    """Return a text string representing the file's contents for LLM context."""
    ext = _ext(filename)
    ct = (content_type or "").lower()

    # ── PDF ───────────────────────────────────────────────────────────────────
    if ext == "pdf" or "pdf" in ct:
        return _extract_pdf(content)

    # ── Word / DOCX ───────────────────────────────────────────────────────────
    if ext in ("docx", "doc"):
        return _extract_docx(content)

    # ── Excel / XLSX / CSV ────────────────────────────────────────────────────
    if ext in ("xlsx", "xls"):
        return _extract_xlsx(content)
    if ext == "csv" or "csv" in ct:
        return _extract_text(content)

    # ── Images → Groq vision (llama-4-scout) ─────────────────────────────────
    if ext in ("jpg", "jpeg", "png", "gif", "webp", "bmp") or ct.startswith("image/"):
        return await _extract_image(filename, content, ct)

    # ── Audio / Video → Groq Whisper ─────────────────────────────────────────
    if ext in ("mp3", "mp4", "wav", "m4a", "ogg", "flac", "webm", "mov", "avi") or \
       ct.startswith("audio/") or ct.startswith("video/"):
        return await _extract_audio(filename, content, ct)

    # ── Plain text / code / JSON / XML / markdown ────────────────────────────
    if ext in ("txt", "md", "json", "xml", "html", "css", "js", "ts", "tsx",
               "jsx", "py", "java", "c", "cpp", "rs", "go", "rb", "sh",
               "yaml", "yml", "toml", "ini", "env") or ct.startswith("text/"):
        return _extract_text(content)

    # ── Generic fallback — attempt UTF-8 decode ───────────────────────────────
    try:
        text = content.decode("utf-8", errors="replace")
        return f"[File: {filename}]\n{text[:_MAX_CHARS]}"
    except Exception:
        return f"[File: {filename} — binary content, {len(content)} bytes — cannot extract text]"


def _extract_pdf(content: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        pages = []
        for i, page in enumerate(reader.pages[:40]):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[Page {i+1}]\n{text.strip()}")
        joined = "\n\n".join(pages)
        return joined[:_MAX_CHARS] if joined else "[PDF — no extractable text]"
    except Exception as e:
        logger.warning("PDF extract failed: %s", e)
        return f"[PDF — extraction failed: {e}]"


def _extract_docx(content: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(content))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return text[:_MAX_CHARS] if text else "[DOCX — no text content]"
    except Exception as e:
        logger.warning("DOCX extract failed: %s", e)
        return f"[DOCX — extraction failed: {e}]"


def _extract_xlsx(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows = []
        for sheet in wb.worksheets[:3]:
            rows.append(f"[Sheet: {sheet.title}]")
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i > 200:
                    break
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(rows)[:_MAX_CHARS]
    except Exception as e:
        logger.warning("XLSX extract failed: %s", e)
        return f"[XLSX — extraction failed: {e}]"


def _extract_text(content: bytes) -> str:
    try:
        import chardet
        enc = chardet.detect(content)["encoding"] or "utf-8"
        text = content.decode(enc, errors="replace")
        return text[:_MAX_CHARS]
    except Exception:
        return content.decode("utf-8", errors="replace")[:_MAX_CHARS]


async def _extract_image(filename: str, content: bytes, content_type: str) -> str:
    """Send image to Groq llama-4-scout vision for description."""
    try:
        import base64
        import httpx
        key = os.environ.get("GROQ_API_KEY", "")
        if not key:
            return f"[Image: {filename} — no vision API key configured]"
        b64 = base64.b64encode(content).decode()
        mime = content_type or "image/jpeg"
        payload = {
            "model": "meta-llama/llama-4-scout-17b-16e-instruct",
            "max_tokens": 1024,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                    {"type": "text", "text": "Describe this image in full detail. Include any text, numbers, names, locations, or other specific information visible. Be thorough."},
                ],
            }],
        }
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
                json=payload,
            )
            if r.status_code == 200:
                desc = r.json()["choices"][0]["message"]["content"]
                return f"[Image: {filename}]\n{desc}"
            logger.warning("Vision API returned %s", r.status_code)
            return f"[Image: {filename} — vision description unavailable (status {r.status_code})]"
    except Exception as e:
        logger.warning("Image extract failed: %s", e)
        return f"[Image: {filename} — vision failed: {e}]"


async def _extract_audio(filename: str, content: bytes, content_type: str) -> str:
    """Transcribe audio/video via Groq Whisper."""
    try:
        import httpx
        key = os.environ.get("GROQ_API_KEY", "")
        if not key:
            return f"[Audio: {filename} — no transcription API key configured]"
        ext = _ext(filename) or "mp3"
        files = {"file": (filename, io.BytesIO(content), content_type or f"audio/{ext}")}
        data = {"model": "whisper-large-v3", "response_format": "text"}
        async with httpx.AsyncClient(timeout=120) as c:
            r = await c.post(
                "https://api.groq.com/openai/v1/audio/transcriptions",
                headers={"Authorization": f"Bearer {key}"},
                files=files,
                data=data,
            )
            if r.status_code == 200:
                return f"[Audio transcript: {filename}]\n{r.text}"
            logger.warning("Whisper returned %s: %s", r.status_code, r.text[:200])
            return f"[Audio: {filename} — transcription failed (status {r.status_code})]"
    except Exception as e:
        logger.warning("Audio extract failed: %s", e)
        return f"[Audio: {filename} — transcription failed: {e}]"
